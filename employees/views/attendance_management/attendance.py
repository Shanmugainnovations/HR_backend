from datetime import datetime, timedelta
import numpy as np
import os
import pytz
from pymongo import MongoClient
import gridfs

IST = pytz.timezone('Asia/Kolkata')

from django.http import JsonResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response

from employees.models import Employee, EmployeeAttendance, SpoofingAttempt
import base64
from employees.face_utils import base64_to_encoding, compare_encodings, imagefile_to_encoding, SpoofingDetectedError, match_face_1_to_n
from pyauth.auth import HasRolePermission

from employees.views.common.utils import to_list, get_mongo_client

# --- 🚀 Performance Cache ---
# Global cache to store employee encodings in memory for faster matching
_ENCODING_CACHE = {
    'matrix': None,      # numpy array of shape (N, 128)
    'employees': [],     # List of employee metadata
    'last_updated': None
}

def get_optimized_encodings(force_refresh=False):
    """
    Returns a numpy matrix of all active employee encodings and their metadata.
    Refreshes automatically or when forced.
    """
    global _ENCODING_CACHE
    now = datetime.now()
    
    # Refresh cache if empty or force_refresh is True
    if force_refresh or _ENCODING_CACHE['matrix'] is None:
        print("🔄 Refreshing Face Encoding Cache...")
        
        # 🚨 Djongo fix: Completely avoid filtering in SQL to prevent SQLDecodeError
        # Since we only have ~300 employees, fetching all and filtering in Python is safe and fast.
        employees = Employee.objects.all()
        
        matrix_list = []
        meta_list = []
        
        for emp in employees:
            # check if active and encoding exists
            if not emp.is_active:
                continue

            # Use the multi-encoding pool (e.g. 3 angles from registration) if present;
            # fall back to the single current_face_encoding for employees registered
            # before this field existed, so nothing existing breaks.
            raw_list = to_list(emp.face_encodings) if emp.face_encodings else None
            if not raw_list:
                raw_list = [emp.current_face_encoding] if emp.current_face_encoding else []

            for raw_enc in raw_list:
                enc = to_list(raw_enc)
                if enc and len(enc) == 128:
                    matrix_list.append(enc)
                    meta_list.append({
                        'employee_id': emp.employee_id,
                        'name': emp.name,
                        'image_md5': emp.image_md5
                    })
        
        if matrix_list:
            _ENCODING_CACHE['matrix'] = np.array(matrix_list)
            _ENCODING_CACHE['employees'] = meta_list
            _ENCODING_CACHE['last_updated'] = now
            print(f"✅ Cache updated: {len(meta_list)} employees loaded.")
        else:
            _ENCODING_CACHE['matrix'] = np.empty((0, 128))
            _ENCODING_CACHE['employees'] = []
            
    return _ENCODING_CACHE['matrix'], _ENCODING_CACHE['employees']

def _save_attendance_record(employee_id, device_id, mode, confidence):
    """
    Internal helper to save attendance data into the database.
    Can be reused by other functions without exposing new URLs.
    """
    return EmployeeAttendance.objects.create(
        employee_id=employee_id,
        device_id=device_id,
        attendence_type=mode,
        confidence=confidence
    )

def _extract_face(img_file, img_b64):
    """Common logic for face extraction."""
    try:
        if img_file:
            return imagefile_to_encoding(img_file)
        elif img_b64:
            return base64_to_encoding(img_b64)
        return None, True
    except Exception as e:
        print(f"🚨 DEBUG: Error extracting face: {e}")
        return None, True

def _match_face(enc):
    """Common logic for finding a face match in the optimized encodings."""
    known_matrix, employee_meta = get_optimized_encodings()
    
    if known_matrix.size == 0:
        return None, 999.0, "No registered employees found"

    meta, dist, err = match_face_1_to_n(enc, known_matrix, employee_meta, threshold=0.45, min_margin=0.05) if enc else (None, 999.0, "No encoding")
    
    if err:
        return meta, dist, err

    return meta, dist, None

def _get_device_label(request):
    """Internal helper to identify the device label using fingerprint."""
    fingerprint = request.headers.get('X-Device-Id') or request.data.get('fingerprint')
    if not fingerprint:
        return "unknown_device"
    
    try:
        client = get_mongo_client()
        if not client: return "unknown_device"
        db_name = os.environ.get("GLOBAL_DB_NAME_HR", "HR")
        db = client[db_name]
        allowed_devices_col = db['employees_alloweddevice']
        
        device_doc = allowed_devices_col.find_one({"fingerprint": fingerprint})
        if device_doc:
            return device_doc.get("label") or "Registered Device"
    except Exception as e:
        print(f"🚨 DEBUG: Error identifying device: {e}")
        
    return "unknown_device"

def _log_spoofing_attempt(employee_id, device_label, img_file, img_b64, category=""):
    """Internal helper to log spoofing attempts."""
    spoofed_image_b64 = ""
    try:
        if img_b64:
            spoofed_image_b64 = img_b64
        elif img_file:
            img_file.seek(0)
            file_content = img_file.read()
            spoofed_image_b64 = base64.b64encode(file_content).decode('utf-8')
    except Exception as e:
        print(f"🚨 DEBUG: Error processing spoofed image for storage: {e}")

    try:
        SpoofingAttempt.objects.create(
            employee_id=employee_id,
            device_id=device_label,
            image=spoofed_image_b64,
            category=category
        )
        print("✅ DEBUG: SpoofingAttempt record created.")
    except Exception as e:
        import traceback
        print(f"🚨 DEBUG: Error creating SpoofingAttempt record: {e}")
        traceback.print_exc()

def _get_registered_image(image_md5):
    """Internal helper to fetch the registered image from MongoDB GridFS."""
    if not image_md5:
        return None
    try:
        client = get_mongo_client()
        if not client: return None
        
        hr_db_name = os.environ.get("GLOBAL_DB_NAME_HR", "HR")
        global_db_name = os.environ.get("GLOBAL_DB_NAME_GLOBAL", "Global")
        
        fs_hr = gridfs.GridFS(client[hr_db_name])
        fs_global = gridfs.GridFS(client[global_db_name])
        
        file_obj = fs_hr.find_one({"md5": image_md5})
        if not file_obj:
            file_obj = fs_global.find_one({"md5": image_md5})
            
        if file_obj:
            img_bytes = file_obj.read()
            return f"data:image/jpeg;base64,{base64.b64encode(img_bytes).decode('utf-8')}"
    except Exception as e:
        print(f"🚨 DEBUG: Error fetching registered image for preview: {e}")
    return None


@api_view(['POST'])
@permission_classes([AllowAny])
def verify_face(request):
    """
    Fast verification endpoint for the first frame in a dual-image sequence.
    Returns success if a confident face match is found, otherwise fails fast.
    """
    image_b64 = request.data.get('image')
    image_file = request.FILES.get('image')
    
    if not image_b64 and not image_file:
        return Response({"error": "Image is required"}, status=400)

    # Identify Device by Fingerprint
    device_label = _get_device_label(request)

    enc, is_real = _extract_face(image_file, image_b64)

    if not enc:
        return Response({"error": "No face found or error processing image"}, status=400)

    matched_meta, best_distance, err = _match_face(enc)
    
    # 🚨 CHECK SPOOFING FIRST, EVEN IF FACE IS UNRECOGNIZED
    if not is_real:
        emp_id = matched_meta['employee_id'] if matched_meta else "unknown"
        print(f"🚨 DEBUG: Spoofing detected (Liveness Check Failed) for {emp_id}!")
        _log_spoofing_attempt(emp_id, device_label, image_file, image_b64, category="SPFV")
        return Response({"error": "Spoofing detected! Real face required."}, status=400)
        
    if err:
        return Response({"error": err}, status=404)
        
    print(f"🏆 VERIFY MATCH: {matched_meta['name']} (Dist: {best_distance:.4f})")
        
    if device_label in ["unknown_device", "Unknown Device"]:
        print(f"🚨 DEBUG: Device spoofing detected for employee {matched_meta['employee_id']}!")
        _log_spoofing_attempt(matched_meta['employee_id'], device_label, image_file, image_b64, category="UNDV")
        return Response({"error": "Unrecognized device. Access denied."}, status=403)
        
    return Response({
        "success": True, 
        "employee_id": matched_meta['employee_id'],
        "name": matched_meta['name'],
        "confidence": best_distance
    }, status=200)

    


@api_view(['POST'])
@permission_classes([AllowAny])
def mark_attendance(request):
    image1_b64 = request.data.get('image')
    image1_file = request.FILES.get('image')
    # Fallback: support frontend sending a single 'image' key
    # if not image1_b64 and not image1_file:
    #     image1_b64 = request.data.get('image')
    #     image1_file = request.FILES.get('image')

    verified_employee_id = request.data.get('verifiedEmployeeID')
    if not verified_employee_id:
        return Response({"error": "verifiedEmployeeID is required"}, status=400)

    print(f"[mark_attendance] Keys received — data: {list(request.data.keys())}, files: {list(request.FILES.keys())}")

    # 0. Identify Device by Fingerprint
    device_label = _get_device_label(request)

    # 1. Extract Encoding & Liveness for image
    enc1, is_real1 = _extract_face(image1_file, image1_b64)

    print(f"[mark_attendance] enc1={'ok' if enc1 else 'EMPTY'}")

    if not enc1:
        return Response({"error": "No face found in image"}, status=400)

    # 2. Find Matching Employee
    meta1, dist1, err1 = _match_face(enc1)
    
    # 🚨 CHECK SPOOFING FIRST, EVEN IF FACE MISMATCHES OR IS UNRECOGNIZED
    if not is_real1:
        emp_id = meta1['employee_id'] if meta1 else verified_employee_id
        print(f"🚨 DEBUG: Spoofing detected (Liveness Check Failed) for employee {emp_id}!")
        _log_spoofing_attempt(emp_id, device_label, image1_file, image1_b64, category="SPFM")
        return Response({"error": "Spoofing detected! Real face required."}, status=400)
    
    if err1:
        print(f"❌ Rejected: {err1}")
        return Response({"error": err1}, status=404)

    if str(meta1['employee_id']) != str(verified_employee_id):
        print(f"❌ Rejected: Face mismatch. Expected {verified_employee_id}, found {meta1['employee_id']}")
        
        # Log the mismatch
        try:
            from employees.models import FaceMismatchLog
            FaceMismatchLog.objects.create(
                verified_employee_id=verified_employee_id,
                mark_employee_id=meta1['employee_id'],
                image=image1_b64 if image1_b64 else "",
                device_id=request.headers.get("X-Device-Id")
            )
        except Exception as e:
            print(f"🚨 Failed to log face mismatch: {e}")
            
        return Response({
            "error": "Face mismatch. Please hold still and try again.",
            "markEmployeeID": meta1['employee_id']
        }, status=400)
        
    best_distance = dist1
    matched_meta = meta1
    is_real = is_real1
        
    print(f"🏆 FINAL WINNER: {matched_meta['name']} (Dist: {best_distance:.4f})")

    # 5. Handle Spoofing (Only if User Found)
    if device_label in ["unknown_device", "Unknown Device"]:
        print(f"🚨 DEBUG: Device spoofing detected for employee {matched_meta['employee_id']}!")
        _log_spoofing_attempt(matched_meta['employee_id'], device_label, image1_file, image1_b64, category="UNDM")
        return Response({"error": "Unrecognized device. Access denied."}, status=403)

    # 6. Mark Attendance (Success)
    mode = request.data.get('mode', 'IN')

    att = _save_attendance_record(
        employee_id=matched_meta['employee_id'],
        device_id=device_label,
        mode=mode,
        confidence=best_distance
    )

    print(f"✅ ATTENDANCE SUCCESS: {matched_meta['name']} (ID: {matched_meta['employee_id']}) | Liveness Verified: {is_real}")

    # Fetch the registered image preview to display on the kiosk
    base64_img = _get_registered_image(matched_meta.get('image_md5'))

    return Response({
        "employee": matched_meta['employee_id'],
        "name": matched_meta['name'],
        "mode": att.attendence_type,
        "timestamp": att.attendence_time.astimezone(IST),
        "confidence": best_distance,
        "registered_image": base64_img
    }, status=201)


@api_view(['GET'])
@permission_classes([AllowAny])
def attendance_report_with_employee_details(request):
    """
    Get attendance records filtered by date and merged with employee details.
    Example: /api/attendance-report/?from_date=2025-10-01&to_date=2025-10-14
    """
    try:
        # ---- Date Filtering ----
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')

        if not from_date or not to_date:
            now = datetime.now()
            from_date = datetime(now.year, now.month, 1)
            to_date = datetime(now.year, now.month + 1, 1) if now.month < 12 else datetime(now.year + 1, 1, 1)
        else:
            from_date = datetime.strptime(from_date, "%Y-%m-%d")
            to_date = datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1)

        # ---- Mongo Connection ----
        db_name = os.environ.get("GLOBAL_DB_NAME", "Global")
        client = get_mongo_client()
        db = client[db_name]

        profiles = db['backend_diagnostics_profile']

        # ---- Fetch Attendance Records (Fast Direct Mongo Query) ----
        # Fetch from previous day to correctly pair night shifts
        from_date_ext = from_date - timedelta(days=1)
        records = list(db['employees_employeeattendance'].find(
            {'attendence_time': {'$gte': from_date_ext, '$lt': to_date}},
            {'_id': 0, 'attendence_id': 1, 'employee_id': 1, 'attendence_time': 1, 'attendence_type': 1, 'device_id': 1}
        ).sort([('employee_id', 1), ('attendence_time', 1)]))

        if not records:
            return Response([], status=200)

        # ---- SQL Department Map ----
        from employees.models import Department
        sql_dept_map = {d.name: d.id for d in Department.objects.all()}

        # ---- Cached Lookup Maps (Fast in-memory) ----
        from employees.views.common.utils import get_cached_reference_maps
        dept_map, desig_map, _ = get_cached_reference_maps()

        # ---- Department Filtering Resolution ----
        department_filter = request.GET.get('department')
        from employees.views.common.utils import resolve_department_filter
        dept_ctx = resolve_department_filter(department_filter)
        profile_query = dept_ctx['mongo_query']

        # ---- Fetch Employee Lookup ----
        employee_map = {}
        # Fetch only employees with face registered from SQL
        face_registered_ids = set(Employee.objects.filter(current_face_encoding__isnull=False).values_list('employee_id', flat=True))

        # Fetch matching profiles from Mongo with projection
        all_profiles = list(profiles.find(profile_query, {'_id': 0, 'employeeId': 1, 'employeeName': 1, 'department': 1, 'designation': 1}))
        for emp in all_profiles:
            emp_id = str(emp.get("employeeId"))
            if emp_id not in face_registered_ids:
                continue
                
            dept_code = emp.get("department")
            dept_name = dept_map.get(dept_code, dept_code)
            employee_map[emp_id] = {
                "employeeName": emp.get("employeeName"),
                "department": dept_name,
                "department_id": sql_dept_map.get(dept_name, dept_code),
                "designation": desig_map.get(emp.get("designation"), emp.get("designation")),
            }
            
        # Fallback for SQL registered employees missing from Global DB
        if not dept_ctx['is_filtered'] or "Unassigned" in dept_ctx['target_terms']:
            registered_sql_emps = Employee.objects.filter(current_face_encoding__isnull=False)
            for sql_emp in registered_sql_emps:
                emp_id_str = str(sql_emp.employee_id)
                if emp_id_str not in employee_map:
                    employee_map[emp_id_str] = {
                        "employeeName": sql_emp.name,
                        "department": "Unassigned",
                        "department_id": None,
                        "designation": "Unassigned"
                    }

        # ---- Prepare Result ----
        result = []

        # Date range for iteration (to_date was already +1 day)
        report_dates = []
        curr = from_date
        while curr < to_date:
            report_dates.append(curr.date())
            curr += timedelta(days=1)

        # Group actual records by (employee_id, date) with Shift Date Logic
        records_by_emp_day = {}
        
        current_emp_id = None
        current_shift_date = None
        last_in_time = None
        
        from datetime import time
        noon_time = time(12, 0)
        
        for r in records:
            raw_time = r.get('attendence_time')
            if not raw_time:
                continue
            ist_time = raw_time.astimezone(IST) if hasattr(raw_time, 'astimezone') else to_ist(raw_time)
            if not ist_time:
                continue
            punch_date = ist_time.date()
            
            r_eid = str(r.get('employee_id', ''))
            if current_emp_id != r_eid:
                current_emp_id = r_eid
                current_shift_date = None
                last_in_time = None
                
            punch_type = r.get('attendence_type')
            assigned_date = punch_date
            
            if punch_type == 'IN':
                current_shift_date = punch_date
                last_in_time = ist_time
                assigned_date = current_shift_date
            elif punch_type == 'OUT':
                if current_shift_date and last_in_time:
                    if (ist_time - last_in_time).total_seconds() <= 16 * 3600:
                        assigned_date = current_shift_date
                    else:
                        if ist_time.time() < noon_time:
                            assigned_date = punch_date - timedelta(days=1)
                else:
                    if ist_time.time() < noon_time:
                        assigned_date = punch_date - timedelta(days=1)
            
            key = (r_eid, assigned_date)
            if key not in records_by_emp_day:
                records_by_emp_day[key] = []
            
            records_by_emp_day[key].append({
                'attendence_id': r.get('attendence_id'),
                'attendence_time': ist_time,
                'attendence_type': punch_type,
                'device_id': r.get('device_id'),
                'employee_id': r_eid
            })

        # Iterate over ALL employees and ALL dates
        # Sorting by name for consistency
        sorted_emp_ids = sorted(employee_map.keys(), key=lambda eid: employee_map[eid]['employeeName'] or "")

        for emp_id in sorted_emp_ids:
            emp_info = employee_map[emp_id]
            # (Additional in-loop filter not needed as employee_map is already filtered)

            for d in report_dates:
                key = (emp_id, d)
                day_records = records_by_emp_day.get(key, [])
                
                if day_records:
                    for r in day_records:
                        result.append({
                            "employee_id": emp_id,
                            "employee_name": emp_info.get("employeeName", "Unknown"),
                            "department": emp_info.get("department", "N/A"),
                            "department_id": emp_info.get("department_id"),
                            "designation": emp_info.get("designation", "N/A"),
                            "device_id": r.get("device_id") if isinstance(r, dict) else getattr(r, 'device_id', 'N/A'),
                            "attendence_type": r.get("attendence_type") if isinstance(r, dict) else getattr(r, 'attendence_type', ''),
                            "attendence_time": r.get("attendence_time") if isinstance(r, dict) else getattr(r, 'attendence_time', None),
                            "confidence": r.get("confidence", 100) if isinstance(r, dict) else getattr(r, 'confidence', 100),
                        })
                else:
                    # No records for this employee on this day -> Add placeholder
                    # Use start of day as placeholder time (localized to IST)
                    placeholder_time = IST.localize(datetime.combine(d, datetime.min.time()))
                    result.append({
                        "employee_id": emp_id,
                        "employee_name": emp_info.get("employeeName", "Unknown"),
                        "department": emp_info.get("department", "N/A"),
                        "department_id": emp_info.get("department_id"),
                        "designation": emp_info.get("designation", "N/A"),
                        "device_id": "N/A",
                        "attendence_type": "ABSENT",
                        "attendence_time": placeholder_time,
                        "confidence": 0,
                    })

        # ---- Export Support ----
        export_mode = request.GET.get('export')
        if export_mode in ['xlsx', 'csv', 'detailed_xlsx']:
            import pandas as pd
            from django.http import HttpResponse

            if export_mode == 'detailed_xlsx':
                from openpyxl import Workbook
                from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Detailed Attendance"
                
                # Header Styling
                header_font = Font(bold=True, color="000000")
                header_fill = PatternFill(start_color="e2e8f0", end_color="e2e8f0", fill_type="solid")
                alignment_center = Alignment(horizontal="center", vertical="center")
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                # Prepare the Matrix Headers
                base_cols = ['S.No', 'Employee ID', 'Employee Name', 'Department', 'Designation']
                # Row 1: Employee info headers (merged vertically) + Dates (merged horizontally 3 cols each)
                # Row 2: Empty under info headers + In/Out/Total under dates
                
                for i, col_name in enumerate(base_cols, start=1):
                    ws.cell(row=1, column=i, value=col_name).font = header_font
                    ws.cell(row=1, column=i).fill = header_fill
                    ws.cell(row=1, column=i).alignment = alignment_center
                    ws.cell(row=1, column=i).border = thin_border
                    ws.merge_cells(start_row=1, start_column=i, end_row=2, end_column=i)

                curr_col = len(base_cols) + 1
                for d in report_dates:
                    date_str = d.strftime('%d/%m/%Y')
                    ws.cell(row=1, column=curr_col, value=date_str).font = header_font
                    ws.cell(row=1, column=curr_col).fill = header_fill
                    ws.cell(row=1, column=curr_col).alignment = alignment_center
                    ws.cell(row=1, column=curr_col).border = thin_border
                    ws.merge_cells(start_row=1, start_column=curr_col, end_row=1, end_column=curr_col + 2)
                    
                    # Row 2 headers
                    for sub_idx, sub_name in enumerate(['In', 'Out', 'Total']):
                        c = curr_col + sub_idx
                        ws.cell(row=2, column=c, value=sub_name).font = header_font
                        ws.cell(row=2, column=c).fill = header_fill
                        ws.cell(row=2, column=c).alignment = alignment_center
                        ws.cell(row=2, column=c).border = thin_border
                    
                    curr_col += 3

                # Data rows
                # Regroup result data by employee and date
                data_map = {} # (emp_id, date) -> {in, out, total}
                # result is a list of all raw punch/absent recs
                
                # To calculate in/out/total, let's group pulses in result
                for r in result:
                    eid = r.get('employee_id')
                    dt_obj = pd.to_datetime(r.get('attendence_time'))
                    d_key = dt_obj.date()
                    if (eid, d_key) not in data_map:
                        data_map[(eid, d_key)] = {'in': None, 'out': None, 'type': r.get('attendence_type')}
                    
                    entry = data_map[(eid, d_key)]
                    p_type = r.get('attendence_type')
                    if p_type == 'IN':
                        if not entry['in'] or dt_obj < pd.to_datetime(entry['in']):
                            entry['in'] = r.get('attendence_time')
                    elif p_type == 'OUT':
                        if not entry['out'] or dt_obj > pd.to_datetime(entry['out']):
                            entry['out'] = r.get('attendence_time')

                # Write employees (sorted)
                sorted_ids = sorted(employee_map.keys())
                row_idx = 3
                for s_idx, eid in enumerate(sorted_ids, 1):
                    info = employee_map[eid]
                    ws.cell(row=row_idx, column=1, value=s_idx).border = thin_border
                    ws.cell(row=row_idx, column=2, value=eid).border = thin_border
                    ws.cell(row=row_idx, column=3, value=info.get('employeeName', 'Unknown')).border = thin_border
                    ws.cell(row=row_idx, column=4, value=info.get('department', 'N/A')).border = thin_border
                    ws.cell(row=row_idx, column=5, value=info.get('designation', 'N/A')).border = thin_border
                    
                    c_idx = 6
                    for d in report_dates:
                        entry = data_map.get((eid, d), {})
                        in_t = "-"
                        out_t = "-"
                        total_h = 0
                        
                        if entry.get('in'):
                            in_t = pd.to_datetime(entry['in']).strftime('%H:%M:%S')
                        if entry.get('out'):
                            out_t = pd.to_datetime(entry['out']).strftime('%H:%M:%S')
                        
                        if entry.get('in') and entry.get('out'):
                            # Calculate hours (Gross)
                            diff = pd.to_datetime(entry['out']) - pd.to_datetime(entry['in'])
                            total_h = round(diff.total_seconds() / 3600, 2)
                            if total_h < 0: total_h = 0
                        
                        ws.cell(row=row_idx, column=c_idx, value=in_t).border = thin_border
                        ws.cell(row=row_idx, column=c_idx+1, value=out_t).border = thin_border
                        ws.cell(row=row_idx, column=c_idx+2, value=total_h).border = thin_border
                        c_idx += 3
                    
                    row_idx += 1

                # Final response
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename="Detailed_Attendance_Report.xlsx"'
                wb.save(response)
                return response

            # Flat Export 
            df = pd.DataFrame(result)
            if not df.empty:
                # Add Day, Month, Year
                df['Timestamp_obj'] = pd.to_datetime(df['attendence_time'])
                df['Day'] = df['Timestamp_obj'].dt.day
                df['Month'] = df['Timestamp_obj'].dt.month
                df['Year'] = df['Timestamp_obj'].dt.year
                
                df = df[['employee_id', 'employee_name', 'department', 'designation', 'Day', 'Month', 'Year', 'attendence_type', 'attendence_time']]
                df.columns = ['Employee ID', 'Name', 'Department', 'Designation', 'Day', 'Month', 'Year', 'Punch Type', 'Timestamp']
                df['Timestamp'] = pd.to_datetime(df['Timestamp']).dt.strftime('%d/%m/%Y %H:%M:%S')
                df.loc[df['Punch Type'] == 'ABSENT', 'Timestamp'] = '-'

            if export_mode == 'xlsx':
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = r'attachment; filename="Attendance_Report.xlsx"'
                df.to_excel(response, index=False, engine='openpyxl')
                return response
            else:
                response = HttpResponse(content_type='text/csv')
                response['Content-Disposition'] = r'attachment; filename="Attendance_Report.csv"'
                df.to_csv(response, index=False)
                return response

        return Response(result, status=200)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)





@api_view(['GET'])
@permission_classes([AllowAny])  # Or admin permission
def get_spoofing_attempts(request):
    """
    Get all recorded spoofing attempts.
    Query Params:
    - month: int (1-12)
    - year: int (e.g. 2025)
    """
    month = request.GET.get('month')
    year = request.GET.get('year')
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')

    queryset = SpoofingAttempt.objects.all()

    if from_date_str and to_date_str:
        try:
            start_date = datetime.strptime(from_date_str, "%Y-%m-%d")
            # For 'to_date', we want to include the entire day, so we go to the start of the next day
            end_date_inclusive = datetime.strptime(to_date_str, "%Y-%m-%d")
            end_date = end_date_inclusive + timedelta(days=1)
            queryset = queryset.filter(timestamp__gte=start_date, timestamp__lt=end_date)
        except (ValueError, TypeError):
            pass
    elif month and year:
        try:
            m = int(month)
            y = int(year)
            start_date = datetime(y, m, 1)
            if m == 12:
                end_date = datetime(y + 1, 1, 1)
            else:
                end_date = datetime(y, m + 1, 1)
            
            queryset = queryset.filter(timestamp__gte=start_date, timestamp__lt=end_date)
        except (ValueError, TypeError):
            pass # Invalid month/year, return all


    department = request.GET.get('department')
    if department and department != 'All':
        try:
            from employees.views.common.utils import resolve_department_filter
            dept_ctx = resolve_department_filter(department)
            dept_emp_ids = dept_ctx['matching_employee_ids'] or set()
            queryset = queryset.filter(employee_id__in=dept_emp_ids)
        except Exception as e:
            print(f"Error filtering spoofing reports by department: {e}")

    attempts = list(queryset.order_by('-timestamp')[:200])
    data = []
    for attempt in attempts:
        data.append({
            "id": attempt.id,
            "employee_id": attempt.employee_id,
            "device_id": attempt.device_id,
            "timestamp": attempt.timestamp.astimezone(IST) if attempt.timestamp else None,
            "image": attempt.image # Base64 string
        })
    return Response(data, status=200)

@api_view(['POST'])
@permission_classes([AllowAny]) # Should be protected in production
def delete_spoofing_attempts(request):
    """
    Delete spoofing attempts by ID.
    Body: { "ids": [1, 2, 3] } or { "ids": "all", "month": 10, "year": 2025 }
    """
    ids = request.data.get('ids')
    
    if ids == "all":
        # Delete all or filtered all
        # (Optional: Add filtering logic for "delete all in this month" if needed)
        # For now, simplistic "delete selected" vs "delete all" logic based on frontend
        # If user wants "delete all in current filter", frontend sends list of all IDs or we handle filter params here.
        # Let's support list of IDs for safety usually.
        # If the user sends "all", we delete EVERYTHING? That's dangerous. 
        # Better to delete everything matching the current filter if filter params provided?
        # Let's stick to list of IDs for explicit deletion unless requested otherwise.
        # Re-reading request: "table type select all and delete" usually implies client-side select all -> send all IDs.
        pass

    if not ids or not isinstance(ids, list):
         return Response({"error": "Invalid IDs provided"}, status=400)

    try:
        SpoofingAttempt.objects.filter(id__in=ids).delete()
        return Response({"message": "Deleted successfully"}, status=200)
    except Exception as e:
        return Response({"error": str(e)}, status=500)
