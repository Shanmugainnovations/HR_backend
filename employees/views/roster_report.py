import csv
import os
import calendar
import pandas as pd
from datetime import date, datetime
from django.http import HttpResponse, JsonResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from pymongo import MongoClient
from ..models import EmployeeShiftSchedule, Employee, Shift, Department

@api_view(['GET'])
@permission_classes([AllowAny])
def export_roster_csv(request):
    from_date_str = request.query_params.get('from_date')
    to_date_str = request.query_params.get('to_date')
    month_str = request.query_params.get('month')
    department_filter = request.query_params.get('department') # Optional (IDs)

    if from_date_str and to_date_str:
        try:
            start_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
            year, month = start_date.year, start_date.month
            month_label = f"{from_date_str}_to_{to_date_str}"
        except ValueError:
            return HttpResponse("Invalid date format. Use YYYY-MM-DD", status=400)
    elif month_str:
        try:
            year, month = map(int, month_str.split('-'))
            _, last_day = calendar.monthrange(year, month)
            start_date = date(year, month, 1)
            end_date = date(year, month, last_day)
            month_label = month_str
        except ValueError:
            return HttpResponse("Invalid month format. Use YYYY-MM", status=400)
    else:
        return HttpResponse("Either from_date/to_date or month parameter is required", status=400)

    # Resolve Department Filter (handle numeric IDs from frontend)
    all_names = []
    codes = []
    raw_ids = []
    
    from ..models import Department
    sql_depts = list(Department.objects.all())
    sql_dept_map = {d.id: d.name for d in sql_depts}

    if department_filter and department_filter != 'All':
        raw_ids = [d.strip() for d in department_filter.split(',')]
        for rid in raw_ids:
            if rid.isdigit():
                d_id = int(rid)
                if d_id in sql_dept_map:
                    all_names.append(sql_dept_map[d_id])
            else:
                all_names.append(rid)

    # 1. Fetch Employees and Department Info from Mongo (Global DB)
    try:
        mongo_uri = os.environ.get("GLOBAL_DB_HOST")
        db_name = os.environ.get("GLOBAL_DB_NAME", "Global")
        client = MongoClient(mongo_uri)
        db = client[db_name]

        profiles_col = db['backend_diagnostics_profile']
        departments_col = db['backend_diagnostics_Departments']
        
        # Create department lookup map
        dept_map = {
            d.get('department_code'): d.get('department_name')
            for d in departments_col.find() # Removed is_active filter
        }

        # Resolve names to codes
        if all_names:
            resolved_codes = list(departments_col.find(
                {"department_name": {"$in": all_names}},
                {"department_code": 1}
            ))
            codes = [c.get("department_code") for c in resolved_codes]

        search_values = all_names + codes + raw_ids

        # Fetch profiles based on department filter
        query = {}
        if search_values:
            query = {
                "$or": [
                    {"department": {"$in": search_values}},
                    {"department_id": {"$in": search_values}},
                    {"department_name": {"$in": search_values}}
                ]
            }
        
        # Fetch only employees with face registered from SQL
        from employees.models import Employee
        face_registered_ids = set(Employee.objects.filter(current_face_encoding__isnull=False).values_list('employee_id', flat=True))

        profiles = list(profiles_col.find(query))

        employees_data = []
        for p in profiles:
            emp_id = str(p.get("employeeId"))
            
            # Skip if not face registered
            if emp_id not in face_registered_ids:
                continue

            dept_code = p.get("department") # This is the ID/Code
            
            employees_data.append({
                "id": emp_id,
                "name": p.get("employeeName"),
                "department": dept_map.get(dept_code, dept_code) or "Unassigned"
            })

        # Sort employees by department then name
        employees_data.sort(key=lambda x: (x['department'], x['name']))

    except Exception as e:
        return HttpResponse(f"Error fetching employee data: {str(e)}", status=500)

    # Date Range defined above

    schedules = EmployeeShiftSchedule.objects.filter(
        date__gte=start_date, 
        date__lte=end_date
    ).select_related('shift', 'employee')

    # Organize schedules: {emp_id: {date_str: shift_name}}
    schedule_map = {}
    date_list = []
    curr = start_date
    while curr <= end_date:
        date_list.append(curr)
        curr += timedelta(days=1)
        if len(date_list) > 62: break

    from datetime import timedelta
    for sch in schedules:
        emp_id = sch.employee.employee_id
        day_str = sch.date.strftime("%Y-%m-%d")
        if emp_id not in schedule_map:
            schedule_map[emp_id] = {}
        
        # Format: ShiftName (HH:MM-HH:MM)
        start_str = sch.shift.start_time.strftime("%H:%M")
        end_str = sch.shift.end_time.strftime("%H:%M")
        schedule_map[emp_id][day_str] = f"{sch.shift.name} ({start_str}-{end_str})"

    # 3. Generate CSV
    filename = f"roster_{month_label}.csv"
    if all_names and len(all_names) == 1:
        filename = f"roster_{month_label}_{all_names[0]}.csv"
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    
    # Header Row
    date_cols = []
    for d_date in date_list:
        day_name = d_date.strftime("%a")
        date_cols.append(f"{d_date.strftime('%d/%m/%Y')} ({day_name})")
        
    header = ["S.No", "Employee ID", "Employee Name", "Department"] + date_cols
    writer.writerow(header)

    # Data Rows
    for idx, emp in enumerate(employees_data, 1):
        row = [idx, emp['id'], emp['name'], emp['department']]
        emp_schedules = schedule_map.get(emp['id'], {})
        
        for d_date in date_list:
            row.append(emp_schedules.get(d_date.strftime("%Y-%m-%d"), "")) # Empty string if no shift
        
        writer.writerow(row)

    return response

@api_view(['GET'])
@permission_classes([AllowAny])
def export_roster_xlsx(request):
    from_date_str = request.query_params.get('from_date')
    to_date_str = request.query_params.get('to_date')
    month_str = request.query_params.get('month')
    department_filter = request.query_params.get('department')

    if from_date_str and to_date_str:
        try:
            start_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
            year, month = start_date.year, start_date.month
            month_label = f"{from_date_str}_to_{to_date_str}"
        except ValueError:
            return HttpResponse("Invalid date format. Use YYYY-MM-DD", status=400)
    elif month_str:
        try:
            year, month = map(int, month_str.split('-'))
            _, last_day = calendar.monthrange(year, month)
            start_date = date(year, month, 1)
            end_date = date(year, month, last_day)
            month_label = month_str
        except ValueError:
            return HttpResponse("Invalid month format. Use YYYY-MM", status=400)
    else:
        return HttpResponse("Either from_date/to_date or month parameter is required", status=400)

    # Resolve Department Info
    from ..models import Department
    sql_depts = list(Department.objects.all())
    sql_dept_map = {d.id: d.name for d in sql_depts}
    all_names = []
    if department_filter and department_filter != 'All':
        raw_ids = [d.strip() for d in department_filter.split(',')]
        for rid in raw_ids:
            if rid.isdigit():
                d_id = int(rid)
                if d_id in sql_dept_map:
                    all_names.append(sql_dept_map[d_id])
            else:
                all_names.append(rid)

    # 1. Fetch Employees from Mongo
    try:
        mongo_uri = os.environ.get("GLOBAL_DB_HOST")
        db_name = os.environ.get("GLOBAL_DB_NAME", "Global")
        client = MongoClient(mongo_uri)
        db = client[db_name]
        profiles_col = db['backend_diagnostics_profile']
        departments_col = db['backend_diagnostics_Departments']
        
        dept_lookup = {d.get('department_code'): d.get('department_name') for d in departments_col.find()}
        
        query = {}
        if all_names:
            resolved_codes = [d.get("department_code") for d in departments_col.find({"department_name": {"$in": all_names}}, {"department_code": 1})]
            search_values = all_names + resolved_codes
            query = {"$or": [{"department": {"$in": search_values}}, {"department_id": {"$in": search_values}}, {"department_name": {"$in": search_values}}]}
        
        # Fetch only employees with face registered from SQL
        from employees.models import Employee
        face_registered_ids = set(Employee.objects.filter(current_face_encoding__isnull=False).values_list('employee_id', flat=True))

        profiles = list(profiles_col.find(query))
        employees_data = []
        for p in profiles:
            eid = str(p.get("employeeId"))
            if eid in face_registered_ids:
                employees_data.append({
                    "id": eid, 
                    "name": p.get("employeeName"), 
                    "department": dept_lookup.get(p.get("department"), p.get("department")) or "Unassigned"
                })
        employees_data.sort(key=lambda x: (x['department'], x['name']))
    except Exception as e:
        return HttpResponse(f"Error fetching data: {str(e)}", status=500)

    # 2. Fetch Schedules
    from datetime import timedelta
    date_list = []
    curr = start_date
    while curr <= end_date:
        date_list.append(curr)
        curr += timedelta(days=1)
        if len(date_list) > 62: break

    schedules = EmployeeShiftSchedule.objects.filter(date__gte=start_date, date__lte=end_date).select_related('shift', 'employee')
    schedule_map = {}
    for sch in schedules:
        emp_id = sch.employee.employee_id
        if emp_id not in schedule_map: schedule_map[emp_id] = {}
        schedule_map[emp_id][sch.date.strftime("%Y-%m-%d")] = sch.shift.name

    # 3. Create DataFrame
    data = []
    for emp in employees_data:
        row = {"Employee ID": emp['id'], "Employee Name": emp['name'], "Department": emp['department']}
        emp_schedules = schedule_map.get(emp['id'], {})
        for d_date in date_list:
            day_name = d_date.strftime("%a")
            col_name = f"{d_date.strftime('%d/%m/%Y')} ({day_name})"
            row[col_name] = emp_schedules.get(d_date.strftime("%Y-%m-%d"), "")
        data.append(row)

    df = pd.DataFrame(data)
    filename = f"roster_{month_label}.xlsx"
    if all_names and len(all_names) == 1:
        filename = f"roster_{month_label}_{all_names[0]}.xlsx"
    
    # 3. Create Excel using openpyxl for basic styling
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Duty Roster"
    
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    sunday_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    sunday_font = Font(color='9C0006', bold=True)

    # 1-Row Header with Full Date
    date_cols = []
    for d_date in date_list:
        day_name = d_date.strftime("%a")
        date_cols.append(f"{d_date.strftime('%d/%m/%Y')} ({day_name})")
        
    header = ["S.No", "Employee ID", "Employee Name", "Department"] + date_cols
    
    # Write Header
    for i, col_name in enumerate(header, 1):
        cell = ws.cell(row=1, column=i, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
        # Color Sundays in header
        if "(" in col_name and "Sun" in col_name:
            cell.fill = sunday_fill
            cell.font = sunday_font

    # Data rows
    for row_idx, emp in enumerate(employees_data, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
        ws.cell(row=row_idx, column=2, value=emp['id']).border = thin_border
        ws.cell(row=row_idx, column=3, value=emp['name']).border = thin_border
        ws.cell(row=row_idx, column=4, value=emp['department']).border = thin_border
        
        emp_schedules = schedule_map.get(emp['id'], {})
        for i, d in enumerate(date_list):
            c = 5 + i
            d_str = d.strftime("%Y-%m-%d")
            cell_val = emp_schedules.get(d_str, "")
            cell = ws.cell(row=row_idx, column=c, value=cell_val)
            cell.border = thin_border
            
            # Highlight Sundays in data rows too
            if d.weekday() == 6:
                cell.fill = sunday_fill

    # Adjust widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    for i in range(5, 5 + len(date_list)):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 15

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="roster_{month_label}.xlsx"'
    wb.save(response)
    return response

@api_view(['POST'])
@permission_classes([AllowAny])
def import_roster_xlsx(request):
    file = request.FILES.get('file')
    from_date_str = request.data.get('from_date')
    to_date_str = request.data.get('to_date')
    month_str = request.data.get('month')
    
    if not file:
        return JsonResponse({"error": "File is required"}, status=400)

    if from_date_str and to_date_str:
        try:
            start_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({"error": "Invalid date format"}, status=400)
    elif month_str:
        try:
            year, month = map(int, month_str.split('-'))
            _, last_day = calendar.monthrange(year, month)
            start_date = date(year, month, 1)
            end_date = date(year, month, last_day)
        except ValueError:
            return JsonResponse({"error": "Invalid month format"}, status=400)
    else:
        return JsonResponse({"error": "Date range or month is required"}, status=400)

    try:
        df = pd.read_excel(file)
        
        if "Employee ID" not in df.columns:
            return JsonResponse({"error": "Missing 'Employee ID' column"}, status=400)

        updated_count = 0
        errors = []
        all_shifts = {s.name.upper(): s for s in Shift.objects.all()}
        
        for _, row in df.iterrows():
            emp_id = str(row['Employee ID']).strip()
            if not emp_id or emp_id == 'nan': continue
            
            try:
                # 🚀 Ensure Employee exists locally (Global Support)
                emp_name = str(row.get('Employee Name', emp_id)).strip()
                if not emp_name or emp_name == 'nan': emp_name = emp_id
                
                employee, _ = Employee.objects.get_or_create(
                    employee_id=emp_id,
                    defaults={'name': emp_name}
                )
            except Exception as e:
                errors.append(f"Error creating employee {emp_id}: {str(e)}")
                continue

            for col in df.columns:
                # 1. Try to parse YYYY-MM-DD from column name
                import re
                target_date = None
                date_match = re.search(r'(\d{4}-\d{2}-\d{2})|(\d{2}/\d{2}/\d{4})', str(col))
                if date_match:
                    try:
                        if date_match.group(1): # YYYY-MM-DD
                            target_date = datetime.strptime(date_match.group(1), '%Y-%m-%d').date()
                        else: # DD/MM/YYYY
                            target_date = datetime.strptime(date_match.group(2), '%d/%m/%Y').date()
                    except: pass
                
                # 2. Fallback to leading digits (day of month) if target_date not found
                if not target_date:
                    match = re.match(r'^(\d+)', str(col))
                    if match:
                        day = int(match.group(1))
                        try:
                            # Use year/month from start_date
                            target_date = date(start_date.year, start_date.month, day)
                        except: pass

                if target_date:
                    try:
                        # Verify it's within our requested range (optional safety)
                        # if target_date < start_date or target_date > end_date: continue
                        
                        val = row[col]
                        shift_name = str(val).strip().upper() if pd.notna(val) else ""
                        
                        if not shift_name:
                            EmployeeShiftSchedule.objects.filter(employee=employee, date=target_date).delete()
                        elif shift_name in all_shifts:
                            EmployeeShiftSchedule.objects.update_or_create(
                                employee=employee,
                                date=target_date,
                                defaults={'shift': all_shifts[shift_name]}
                            )
                            updated_count += 1
                        else:
                            errors.append(f"Invalid shift {shift_name} for {emp_id} on day {day}")
                    except Exception as e:
                        errors.append(f"Error on day {day} for {emp_id}: {str(e)}")


        return JsonResponse({"message": f"Updated {updated_count} shifts", "errors": errors[:10]})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@api_view(['POST'])
@permission_classes([AllowAny])
def preview_roster_xlsx(request):
    file = request.FILES.get('file')
    from_date_str = request.data.get('from_date')
    to_date_str = request.data.get('to_date')
    month_str = request.data.get('month')
    
    if not file:
        return JsonResponse({"error": "File is required"}, status=400)

    if from_date_str and to_date_str:
        try:
            start_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({"error": "Invalid date format"}, status=400)
    elif month_str:
        try:
            year, month = map(int, month_str.split('-'))
            _, last_day = calendar.monthrange(year, month)
            start_date = date(year, month, 1)
            end_date = date(year, month, last_day)
        except ValueError:
            return JsonResponse({"error": "Invalid month format"}, status=400)
    else:
        return JsonResponse({"error": "Date range or month is required"}, status=400)

    try:
        df = pd.read_excel(file)
        
        if "Employee ID" not in df.columns:
            return JsonResponse({"error": "Missing 'Employee ID' column"}, status=400)

        preview_data = []
        errors = []
        all_shifts = {s.name.upper(): s.name for s in Shift.objects.all()}
        
        # Mapping of department name to set of valid shift names
        dept_shift_map = {}
        for d in Department.objects.prefetch_related('shifts').all():
            dept_shift_map[d.name.upper()] = {s.name.upper() for s in d.shifts.all()}

        detected_dates = set()

        for _, row in df.iterrows():
            emp_id = str(row['Employee ID']).strip()
            if not emp_id or emp_id == 'nan': continue
            
            emp_name = str(row.get('Employee Name', emp_id)).strip()
            if not emp_name or emp_name == 'nan': emp_name = emp_id
            
            dept_name = str(row.get('Department', '')).strip()
            dept_name_upper = dept_name.upper()
            
            valid_shifts_for_dept = dept_shift_map.get(dept_name_upper, set())
            
            emp_preview = {
                "id": emp_id,
                "name": emp_name,
                "department": dept_name,
                "shifts": {}
            }
            
            for col in df.columns:
                import re
                target_date = None
                date_match = re.search(r'(\d{4}-\d{2}-\d{2})|(\d{2}/\d{2}/\d{4})', str(col))
                if date_match:
                    try:
                        if date_match.group(1): # YYYY-MM-DD
                            target_date = datetime.strptime(date_match.group(1), '%Y-%m-%d').date()
                        else: # DD/MM/YYYY
                            target_date = datetime.strptime(date_match.group(2), '%d/%m/%Y').date()
                    except: pass
                
                if not target_date:
                    match = re.match(r'^(\d+)', str(col))
                    if match:
                        day = int(match.group(1))
                        try:
                            target_date = date(start_date.year, start_date.month, day)
                        except: pass

                if target_date:
                    date_str = target_date.strftime("%Y-%m-%d")
                    detected_dates.add(date_str)
                    try:
                        val = row[col]
                        shift_name = str(val).strip().upper() if pd.notna(val) else ""
                        
                        if shift_name:
                            # Global existence check
                            exists_globally = shift_name in all_shifts
                            
                            # Department-based check
                            # If the department is not found in shift config, we only use global check
                            # But if the department IS configured, we must match its shifts.
                            is_valid = exists_globally
                            if dept_name_upper in dept_shift_map:
                                is_valid = shift_name in valid_shifts_for_dept
                            
                            emp_preview["shifts"][date_str] = {
                                "name": all_shifts.get(shift_name, str(val).strip()),
                                "is_valid": is_valid
                            }
                            if not is_valid:
                                if not exists_globally:
                                    errors.append(f"Shift '{shift_name}' not found in global config.")
                                else:
                                    errors.append(f"Shift '{shift_name}' is not configured for department '{dept_name}'.")
                    except Exception as e:
                        errors.append(f"Error on {target_date} for {emp_id}: {str(e)}")
            
            preview_data.append(emp_preview)

        sorted_headers = sorted(list(detected_dates))
        return JsonResponse({
            "preview": preview_data,
            "headers": sorted_headers,
            "errors": errors[:20],
            "total_employees": len(preview_data),
        })
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
@api_view(['POST'])
@permission_classes([AllowAny])
def approve_roster_data(request):
    data = request.data.get('preview')
    if not data:
        return JsonResponse({"error": "No data found"}, status=400)
    
    all_shifts = {s.name.upper(): s for s in Shift.objects.all()}
    updated_count = 0
    errors = []

    for emp_data in data:
        emp_id = emp_data.get('id')
        emp_name = emp_data.get('name')
        shifts = emp_data.get('shifts', {})

        if not emp_id: continue
        
        try:
            employee, _ = Employee.objects.get_or_create(
                employee_id=emp_id,
                defaults={'name': emp_name or emp_id}
            )
        except Exception as e:
            errors.append(f"Error creating/fetching employee {emp_id}: {str(e)}")
            continue

        for date_str, shift_info in shifts.items():
            shift_name = shift_info.get('name', '').strip().upper()
            try:
                target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                if not shift_name:
                    EmployeeShiftSchedule.objects.filter(employee=employee, date=target_date).delete()
                elif shift_name in all_shifts:
                    EmployeeShiftSchedule.objects.update_or_create(
                        employee=employee,
                        date=target_date,
                        defaults={'shift': all_shifts[shift_name]}
                    )
                    updated_count += 1
                else:
                    errors.append(f"Shift '{shift_name}' not found globally for employee {emp_id}")
            except Exception as e:
                errors.append(f"Error for {emp_id} on {date_str}: {str(e)}")

    return JsonResponse({"message": f"Successfully updated {updated_count} shifts", "errors": errors[:10]})
