from datetime import date, datetime, timedelta
import calendar
import os
import pytz
from pymongo import MongoClient
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from django.db.models import Min, Max, F
from django.db.models.expressions import RawSQL
from ..models import EmployeeShiftSchedule, EmployeeAttendance, Shift

IST = pytz.timezone('Asia/Kolkata')
# 🔥 Safe conversion function (IMPORTANT)
def to_ist(dt):
    if not dt:
        return None

    # If string → convert to datetime
    if isinstance(dt, str):
        try:
            dt = datetime.fromisoformat(dt)
        except:
            return None

    # If naive datetime → make UTC
    if dt.tzinfo is None:
        dt = pytz.UTC.localize(dt)

    return dt.astimezone(IST)

@api_view(['GET'])
@permission_classes([AllowAny])
def roster_attendance_report(request):
    """
    Combines Shift Roster with Actual Punch timings and Total Hours.
    """
    month_str = request.query_params.get('month')
    from_date_str = request.query_params.get('from_date')
    to_date_str = request.query_params.get('to_date')
    department_filter = request.query_params.get('department')
    
    if not month_str and not (from_date_str and to_date_str):
        return Response({"error": "Month parameter (YYYY-MM) or From-To dates are required"}, status=400)

    # Resolve Department Filter (handle numeric IDs from frontend)
    all_names = []
    codes = []
    raw_ids = []
    
    from ..models import Department
    sql_depts = list(Department.objects.all())
    sql_dept_map = {d.id: d.name for d in sql_depts}
    name_to_sql_id = {d.name: d.id for d in sql_depts}

    if department_filter and department_filter != 'All':
        raw_ids = [d.strip() for d in department_filter.split(',')]
        for rid in raw_ids:
            if rid.isdigit():
                d_id = int(rid)
                if d_id in sql_dept_map:
                    all_names.append(sql_dept_map[d_id])
            else:
                all_names.append(rid)

    # 1. Fetch Employees and Department Info from Mongo (Global DB)
    try:
        mongo_uri = os.environ.get("GLOBAL_DB_HOST")
        db_name = os.environ.get("GLOBAL_DB_NAME", "Global")
        client = MongoClient(mongo_uri)
        db = client[db_name]

        profiles_col = db['backend_diagnostics_profile']
        departments_col = db['backend_diagnostics_Departments']
        designations_col = db['backend_diagnostics_Designation']
        
        # Create lookup maps
        dept_map = {
            d.get('department_code'): d.get('department_name')
            for d in departments_col.find() # Removed is_active=True to be safe
        }
        
        # Resolve names to codes
        if all_names:
            resolved_codes = list(departments_col.find(
                {"department_name": {"$in": all_names}},
                {"department_code": 1}
            ))
            codes = [c.get("department_code") for c in resolved_codes]

        search_values = all_names + codes + raw_ids

        desig_map = {
            d.get('Designation_code'): d.get('designation')
            for d in designations_col.find({'is_active': True})
        }

        # Fetch profiles based on department filter
        query = {}
        if search_values:
            query = {
                "$or": [
                    {"department": {"$in": search_values}},
                    {"department_id": {"$in": search_values}},
                    {"department_name": {"$in": search_values}}
                ]
            }
        
        # Fetch all employees from SQL
        from employees.models import Employee
        active_employees = Employee.objects.all()
        active_employee_ids = set(active_employees.values_list('employee_id', flat=True))

        profiles = list(profiles_col.find(query))

        employees_data = {} # {emp_id: {name, department, department_id, designation}}
        for p in profiles:
            emp_id = str(p.get("employeeId"))
            
            # Skip if not in SQL DB
            if emp_id not in active_employee_ids:
                continue

            dept_code = p.get("department") # This is the ID/Code
            dept_name = dept_map.get(dept_code, dept_code) or "Unassigned"
            
            # Use SQL ID for department_id in response if possible
            sql_id = name_to_sql_id.get(dept_name, dept_code)

            employees_data[emp_id] = {
                "name": p.get("employeeName"),
                "department": dept_name,
                "department_id": sql_id,
                "designation": desig_map.get(p.get("designation"), p.get("designation")) or "Unassigned"
            }
            
        # Fallback: Add employees that are in SQL but not in Global DB
        # Only add them if there is no department filter, or if "Unassigned" is explicitly requested.
        if not search_values or "Unassigned" in search_values:
            for emp in active_employees:
                emp_id_str = str(emp.employee_id)
                if emp_id_str not in employees_data:
                    employees_data[emp_id_str] = {
                        "name": emp.name,
                        "department": "Unassigned",
                        "department_id": None,
                        "designation": "Unassigned"
                    }
        
    except Exception as e:
        return Response({"error": f"Error fetching employee data: {str(e)}"}, status=500)

    if not employees_data:
        return Response([], status=200)

    # 2. Date Range
    if from_date_str and to_date_str:
        try:
            start_date = datetime.strptime(from_date_str, "%Y-%m-%d").date()
            end_date = datetime.strptime(to_date_str, "%Y-%m-%d").date()
        except ValueError:
            return Response({"error": "Invalid date format. Use YYYY-MM-DD"}, status=400)
    elif month_str:
        try:
            year, month = map(int, month_str.split('-'))
            _, last_day = calendar.monthrange(year, month)
            start_date = date(year, month, 1)
            end_date = date(year, month, last_day)
        except (ValueError, TypeError):
             return Response({"error": "Invalid month format. Use YYYY-MM"}, status=400)
    else:
        return Response({"error": "Date range or month is required"}, status=400)

    # Pre-calculate dates for report
    report_dates = []
    curr = start_date
    while curr <= end_date:
        report_dates.append(curr)
        curr += timedelta(days=1)

    # 3. Fetch Shift Schedules
    schedules = EmployeeShiftSchedule.objects.filter(
        date__gte=start_date, 
        date__lte=end_date,
        employee_id__in=employees_data.keys()
    ).select_related('shift', 'employee')

    schedule_map = {} # {(emp_id, date): shift_obj}
    for sch in schedules:
        schedule_map[(sch.employee_id, sch.date)] = sch.shift

    # 4. Fetch Attendance (Grouped by Date & Employee)
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())
    
    # Using 'attendence_time' and 'employee_id' as per model
    attendance_records = EmployeeAttendance.objects.filter(
        attendence_time__range=(start_dt, end_dt),
        employee_id__in=employees_data.keys()
    ).order_by('employee_id', 'attendence_time')

    # Process attendance: map (emp_id, date) -> list of punches
    attendance_map = {}
    for att in attendance_records:
        # Convert to IST for reporting
        ist_time = to_ist(att.attendence_time)
        d_date = ist_time.date()
        key = (att.employee_id, d_date)
        
        if key not in attendance_map:
            attendance_map[key] = []
        
        attendance_map[key].append({'time': ist_time, 'type': att.attendence_type})

    # 5. Build Final Report Data
    report_data = []
    
    # Sorting employees by ID for display
    import re
    def natural_sort_key(s):
        return [int(text) if text.isdigit() else text.lower() for text in re.split('([0-9]+)', str(s))]

    sorted_emp_ids = sorted(employees_data.keys(), key=natural_sort_key)

    for emp_id in sorted_emp_ids:
        emp_info = employees_data[emp_id]
        
        for current_date in report_dates:
            
            key = (emp_id, current_date)
            # Tuple key for shift lookup: (emp_id, date) 
            # Check how schedule_map was built: ((sch.employee.employee_id, sch.date))? No, sch.employee_id is charfield?
            # Model: employee = ForeignKey(Employee). Employee model has primary key 'employee_id'.
            # So sch.employee_id returns the string ID.
            
            shift_obj = schedule_map.get((emp_id, current_date))
            punches = attendance_map.get((emp_id, current_date), [])
            
            # Default values
            shift_name = "Off/Unassigned"
            shift_timing = "-"
            check_in_str = "-"
            check_out_str = "-"
            total_hours_str = "-"
            late_early_hrs = "-"
            status = "Absent"

            if shift_obj:
                shift_name = shift_obj.name
                shift_timing = f"{shift_obj.start_time.strftime('%H:%M')} - {shift_obj.end_time.strftime('%H:%M')}"

            is_leave_shift = False
            if shift_obj:
                is_leave_shift = (
                    (shift_obj.start_time.strftime('%H:%M') == '00:00' and shift_obj.end_time.strftime('%H:%M') == '00:00')
                    or shift_name.upper() in ['OFF', 'EL', 'CL', 'SL', 'ML', 'COFF', 'LEAVE', 'WEEK OFF', 'PH', 'COL']
                )

            check_in_dt = None
            check_out_dt = None

            if punches:
                in_punches = [p['time'] for p in punches if p['type'] == 'IN']
                out_punches = [p['time'] for p in punches if p['type'] == 'OUT']
                
                check_in_dt = in_punches[0] if in_punches else None
                check_out_dt = out_punches[-1] if out_punches else None
                
                if check_in_dt:
                    check_in_str = check_in_dt.strftime('%H:%M:%S')
                
                if not shift_obj:
                    status = "UA"
                elif is_leave_shift:
                    status = "W/O"
                else:
                    status = "Present" # Basic logic

                if check_in_dt and check_out_dt and check_out_dt > check_in_dt:
                    check_out_str = check_out_dt.strftime('%H:%M:%S')
                    duration = check_out_dt - check_in_dt
                    total_seconds = duration.total_seconds()
                    hours = int(total_seconds // 3600)
                    minutes = int((total_seconds % 3600) // 60)
                    total_hours_str = f"{hours}h {minutes}m"

                    if status in ["UA", "W/O"] and total_seconds >= 28800:
                        status = f"P({status})"
                else:
                    # Only one punch or missing OUT
                    if check_out_dt and not check_in_dt:
                        check_out_str = check_out_dt.strftime('%H:%M:%S')
                        
                    if status == "Present":
                        status = "Single Punch"
                    total_hours_str = "0h 0m"

            # Check if Absent (Shift assigned but no punches)
            if shift_obj and not punches:
                if is_leave_shift:
                    status = shift_name
                else:
                    status = "Absent"
            # Check if Week Off (No shift assigned)
            elif not shift_obj and not punches:
                status = "Week Off/Holiday"
            
            # Refine status based on shift timings if present and employee was present
            if status == "Present" and shift_obj and check_in_dt and check_out_dt:
                try:
                    shift_start = shift_obj.start_time
                    shift_end = shift_obj.end_time
                    
                    first_punch_time = check_in_dt.time()
                    last_punch_time = check_out_dt.time()

                    # Late Login Logic (e.g., > 15 mins grace? defaulting to strict for now or 5 mins)
                    # Adding 10 mins grace for now
                    # Need to combine with dummy date to do arithmetic
                    
                    # Convert to datetime to add timedelta
                    dummy_date = date(2000, 1, 1)
                    shift_start_dt = datetime.combine(dummy_date, shift_start)
                    punch_in_dt = datetime.combine(dummy_date, first_punch_time)
                    
                    shift_end_dt = datetime.combine(dummy_date, shift_end)
                    if shift_end < shift_start:
                        shift_end_dt += timedelta(days=1)
                    
                    punch_out_dt = datetime.combine(dummy_date, last_punch_time)
                    if last_punch_time < shift_start:
                        punch_out_dt += timedelta(days=1)
                        
                    is_late_login = (punch_in_dt - shift_start_dt).total_seconds() > 600
                    is_early_checkout = (shift_end_dt - punch_out_dt).total_seconds() > 600

                    if is_late_login and is_early_checkout:
                        status = "Late In & Early Out"
                    elif is_late_login:
                        status = "Late Login"
                    elif is_early_checkout:
                        status = "EG"
                    else:
                        status = "Present"

                except Exception as e:
                    # Keep as Present if error
                    pass

            report_data.append({
                "date": current_date.strftime("%Y-%m-%d"),
                "employee_id": emp_id,
                "employee_name": emp_info['name'],
                "department": emp_info['department'],
                "department_id": emp_info['department_id'],
                "designation": emp_info['designation'],
                "shift_name": shift_name,
                "shift_timing": shift_timing,
                "check_in": check_in_str,
                "check_out": check_out_str,
                "total_hours": total_hours_str,
                "late_early_hrs": late_early_hrs,
                "status": status
            })

    # 6. Export Support
    export_format = request.query_params.get('export')
    
    if export_format == 'xlsx':
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from django.http import HttpResponse

        wb = Workbook()
        ws = wb.active
        ws.title = "Roster Attendance"

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Base Columns
        base_cols = ["S.No", "Employee ID", "Employee Name", "Department", "Designation"]
        for i, col in enumerate(base_cols, 1):
            cell = ws.cell(row=1, column=i, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            ws.merge_cells(start_row=1, start_column=i, end_row=2, end_column=i)

        # Date Columns (5 per date: Shift, In, Out, Total, Late/Early)
        curr_col = len(base_cols) + 1
        for d in report_dates:
            d_name = d.strftime("%a")
            # header date label
            date_label = f"{d.strftime('%d/%m/%Y')} ({d_name})"
            
            cell = ws.cell(row=1, column=curr_col, value=date_label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            ws.merge_cells(start_row=1, start_column=curr_col, end_row=1, end_column=curr_col + 4)
            
            subs = ["Shift", "In", "Out", "Total", "Late/Early"]
            for j, sub in enumerate(subs):
                c = curr_col + j
                sc = ws.cell(row=2, column=c, value=sub)
                sc.font = header_font
                sc.fill = header_fill
                sc.alignment = center_align
                sc.border = thin_border
            
            curr_col += 5

        # Data rows
        from collections import defaultdict
        grouped_data = defaultdict(dict)
        for item in report_data:
            grouped_data[item['employee_id']][item['date']] = item

        row_idx = 3
        for s_no, eid in enumerate(sorted_emp_ids, 1):
            emp_info = employees_data[eid]
            # Info
            ws.cell(row=row_idx, column=1, value=s_no).border = thin_border
            ws.cell(row=row_idx, column=2, value=eid).border = thin_border
            ws.cell(row=row_idx, column=3, value=emp_info['name']).border = thin_border
            ws.cell(row=row_idx, column=4, value=emp_info['department']).border = thin_border
            ws.cell(row=row_idx, column=5, value=emp_info['designation']).border = thin_border
            
            c_idx = 6
            for d in report_dates:
                d_str = d.strftime("%Y-%m-%d")
                d_item = grouped_data[eid].get(d_str, {})
                
                s_name = d_item.get('shift_name', '-')
                if s_name == "Off/Unassigned": s_name = "O(-)"
                
                ws.cell(row=row_idx, column=c_idx, value=s_name).border = thin_border
                ws.cell(row=row_idx, column=c_idx+1, value=d_item.get('check_in', '-')).border = thin_border
                ws.cell(row=row_idx, column=c_idx+2, value=d_item.get('check_out', '-')).border = thin_border
                ws.cell(row=row_idx, column=c_idx+3, value=d_item.get('total_hours', '-')).border = thin_border
                ws.cell(row=row_idx, column=c_idx+4, value=d_item.get('late_early_hrs', '-')).border = thin_border
                c_idx += 5
            row_idx += 1

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Roster_Attendance_Report.xlsx"'
        wb.save(response)
        return response

    if export_format == 'flat_xlsx':
        import pandas as pd
        from django.http import HttpResponse

        df = pd.DataFrame(report_data)
        if not df.empty:
            # Convert date column format
            df['date'] = df['date'].apply(lambda x: datetime.strptime(x, "%Y-%m-%d").strftime("%d/%m/%Y"))
            df['Day'] = df['date'].apply(lambda x: datetime.strptime(x, "%d/%m/%Y").day)
            df['Month'] = df['date'].apply(lambda x: datetime.strptime(x, "%d/%m/%Y").strftime("%B"))
            df['Year'] = df['date'].apply(lambda x: datetime.strptime(x, "%d/%m/%Y").year)
            
            # Reorder
            df = df[['date', 'Day', 'Month', 'Year', 'employee_id', 'employee_name', 'department', 'designation', 'shift_name', 'shift_timing', 'check_in', 'check_out', 'total_hours', 'late_early_hrs', 'status']]
            # Rename for display
            df.columns = ['Date', 'Day', 'Month', 'Year', 'Employee ID', 'Employee Name', 'Department', 'Designation', 'Allocated Shift', 'Shift Timings', 'Check In', 'Check Out', 'Total Hours', 'Late / Early', 'Status']

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = r'attachment; filename="Roster_List_Report.xlsx"'
        df.to_excel(response, index=False, engine='openpyxl')
        return response

    if export_format == 'csv':
        import csv
        from django.http import HttpResponse

        response = HttpResponse(content_type='text/csv')
        period = f"{from_date_str}_to_{to_date_str}" if from_date_str else str(month_str)
        filename = f"Roster_Actual_Report_{period}.csv"
        if department_filter:
            filename = f"Roster_Actual_Report_{period}_{department_filter}.csv"
            
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)
        writer.writerow(['Date', 'Day', 'Month', 'Year', 'Employee ID', 'Employee Name', 'Department', 'Designation', 'Allocated Shift', 'Shift Timings', 'Check In', 'Check Out', 'Total Hours', 'Late / Early', 'Status'])

        for row in report_data:
            d_dt = datetime.strptime(row['date'], "%Y-%m-%d")
            writer.writerow([
                d_dt.strftime("%d/%m/%Y"), d_dt.day, d_dt.strftime("%B"), d_dt.year,
                row['employee_id'], row['employee_name'], row['department'], row['designation'],
                row['shift_name'], row['shift_timing'], row['check_in'], row['check_out'], row['total_hours'],
                row['late_early_hrs'], row['status']
            ])
        
        return response

    if export_format == 'summary_xlsx':
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from django.http import HttpResponse
        from collections import defaultdict

        wb = Workbook()
        ws = wb.active
        ws.title = "Summary Attendance"

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Base Columns
        base_cols = ["S.No", "Employee ID", "Employee Name", "Department", "Designation"]
        
        # Header Row
        for i, col in enumerate(base_cols, 1):
            cell = ws.cell(row=1, column=i, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Date Columns
        for i, d in enumerate(report_dates, len(base_cols) + 1):
            date_label = d.strftime('%d/%m/%Y')
            cell = ws.cell(row=1, column=i, value=date_label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Group data
        grouped_data = defaultdict(dict)
        for item in report_data:
            grouped_data[item['employee_id']][item['date']] = item

        # Data rows
        row_idx = 2
        for s_no, eid in enumerate(sorted_emp_ids, 1):
            emp_info = employees_data[eid]
            ws.cell(row=row_idx, column=1, value=s_no).border = thin_border
            ws.cell(row=row_idx, column=2, value=eid).border = thin_border
            ws.cell(row=row_idx, column=3, value=emp_info['name']).border = thin_border
            ws.cell(row=row_idx, column=4, value=emp_info['department']).border = thin_border
            ws.cell(row=row_idx, column=5, value=emp_info['designation']).border = thin_border
            
            for i, d in enumerate(report_dates, len(base_cols) + 1):
                d_str = d.strftime("%Y-%m-%d")
                d_item = grouped_data[eid].get(d_str, {})
                status = d_item.get('status', 'Absent')
                
                # Determine abbreviation
                abbr = '-'
                if status == 'Present' or status == 'Mismatched Punch':
                    abbr = 'P'
                elif status == 'Late Login':
                    abbr = 'P(LL)'
                elif status == 'Late In & Early Out':
                    abbr = 'LI/EO'
                elif status == 'Early Checkout':
                    abbr = 'EC'
                elif status == 'Single Punch':
                    abbr = 'SP'
                elif status == 'EG':
                    abbr = 'EG'
                elif status == 'Absent' or status == 'Week Off/Holiday':
                    # User requested: "illanan A nu kaatu" -> If not explicitly assigned an OFF/Leave, and no punch, it's A
                    abbr = 'A'
                elif status == 'WF':
                    abbr = 'WF'
                elif status == 'UA':
                    abbr = 'UA'
                elif status == 'P(UA)':
                    abbr = 'P(UA)'
                elif status == 'P(W/O)':
                    abbr = 'P(W/O)'
                else:
                    # Should be EL, CL, SL, OFF, etc from our previous fix
                    abbr = status

                cell = ws.cell(row=row_idx, column=i, value=abbr)
                cell.border = thin_border
                cell.alignment = center_align
                
                # Colors
                if abbr in ['P', 'P(UA)', 'P(W/O)']:
                    cell.font = Font(color="10b981", bold=True)
                elif abbr == 'A':
                    cell.font = Font(color="ef4444", bold=True)
                elif abbr in ['EG', 'SP', 'P(LL)']:
                    cell.font = Font(color="f59e0b", bold=True) # Amber color

                elif abbr == 'WF':
                    cell.font = Font(color="8b5cf6", bold=True) # Purple color
                elif abbr == 'UA':
                    cell.font = Font(color="db2777", bold=True) # Pink color
                else:
                    cell.font = Font(color="3b82f6", bold=True)

            row_idx += 1

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Roster_Summary_Report.xlsx"'
        wb.save(response)
        return response

    return Response(report_data)
