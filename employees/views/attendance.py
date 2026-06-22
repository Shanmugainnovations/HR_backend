from datetime import datetime, timedelta
import numpy as np
import os
import pytz
from pymongo import MongoClient
import gridfs

IST = pytz.timezone('Asia/Kolkata')

from django.http import JsonResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response

from employees.models import Employee, EmployeeAttendance, SpoofingAttempt
import base64
from employees.face_utils import base64_to_encoding, compare_encodings, imagefile_to_encoding, SpoofingDetectedError
from pyauth.auth import HasRolePermission

from .utils import to_list

# --- 🚀 Performance Cache ---
# Global cache to store employee encodings in memory for faster matching
_ENCODING_CACHE = {
    'matrix': None,      # numpy array of shape (N, 128)
    'employees': [],     # List of employee metadata
    'last_updated': None
}

def get_optimized_encodings(force_refresh=False):
    """
    Returns a numpy matrix of all active employee encodings and their metadata.
    Refreshes automatically or when forced.
    """
    global _ENCODING_CACHE
    now = datetime.now()
    
    # Refresh cache if empty or force_refresh is True
    if force_refresh or _ENCODING_CACHE['matrix'] is None:
        print("🔄 Refreshing Face Encoding Cache...")
        
        # 🚨 Djongo fix: Completely avoid filtering in SQL to prevent SQLDecodeError
        # Since we only have ~300 employees, fetching all and filtering in Python is safe and fast.
        employees = Employee.objects.all()
        
        matrix_list = []
        meta_list = []
        
        for emp in employees:
            # check if active and encoding exists
            if not emp.is_active:
                continue
                
            raw_enc = emp.current_face_encoding
            if not raw_enc:
                continue
                
            enc = to_list(raw_enc)
            if enc and len(enc) == 128:
                matrix_list.append(enc)
                meta_list.append({
                    'employee_id': emp.employee_id,
                    'name': emp.name
                })
        
        if matrix_list:
            _ENCODING_CACHE['matrix'] = np.array(matrix_list)
            _ENCODING_CACHE['employees'] = meta_list
            _ENCODING_CACHE['last_updated'] = now
            print(f"✅ Cache updated: {len(meta_list)} employees loaded.")
        else:
            _ENCODING_CACHE['matrix'] = np.empty((0, 128))
            _ENCODING_CACHE['employees'] = []
            
    return _ENCODING_CACHE['matrix'], _ENCODING_CACHE['employees']

@api_view(['POST'])
# @permission_classes([HasRolePermission])
def mark_attendance(request):
    image1_b64 = request.data.get('image1')
    image2_b64 = request.data.get('image2')
    image1_file = request.FILES.get('image1')
    image2_file = request.FILES.get('image2')
    
    # Fallback for old single-image payload
    if not image1_b64 and not image1_file:
        image1_b64 = request.data.get('image')
        image1_file = request.FILES.get('image')

    employee_id = request.data.get('auth-user-id')

    # 0. Identify Device by Fingerprint
    fingerprint = request.headers.get('X-Device-Id') or request.data.get('fingerprint')
    device_label = "unknown_device"
    
    if fingerprint:
        try:
            mongo_uri = os.environ.get("GLOBAL_DB_HOST")
            db_name = os.environ.get("GLOBAL_DB_NAME_HR", "HR") # Consistent with auth.py
            client = MongoClient(mongo_uri)
            db = client[db_name]
            allowed_devices_col = db['employees_alloweddevice']
            
            device_doc = allowed_devices_col.find_one({"fingerprint": fingerprint})
            if device_doc:
                device_label = device_doc.get("label", "Unknown Device")
        except Exception as e:
            print(f"🚨 DEBUG: Error identifying device: {e}")

    def extract_face(img_file, img_b64):
        try:
            if img_file:
                return imagefile_to_encoding(img_file)
            elif img_b64:
                return base64_to_encoding(img_b64)
            return None, True
        except Exception as e:
            print(f"🚨 DEBUG: Error extracting face: {e}")
            return None, True

    # 1. Extract Encoding & Liveness for both images
    enc1, is_real1 = extract_face(image1_file, image1_b64)
    enc2, is_real2 = extract_face(image2_file, image2_b64)

    if not enc1 and not enc2:
        return Response({"error": "No face found in images"}, status=400)

    # 2. Find Matching Employee (Vectorized Optimization)
    known_matrix, employee_meta = get_optimized_encodings()
    
    if known_matrix.size == 0:
        return Response({"error": "No registered employees found"}, status=404)

    def get_best_match(enc):
        if not enc:
            return None, 999.0
        enc_np = np.array(enc)
        distances = np.linalg.norm(known_matrix - enc_np, axis=1)
        best_idx = np.argmin(distances)
        return employee_meta[best_idx], float(distances[best_idx])

    meta1, dist1 = get_best_match(enc1)
    meta2, dist2 = get_best_match(enc2)
    
    MATCH_THRESHOLD = 0.45
    best_distance = dist1
    matched_meta = meta1
    is_real = is_real1
    
    if meta1 and meta2:
        if meta1['employee_id'] != meta2['employee_id']:
            print(f"❌ Rejected: Inconsistent match ({meta1['name']} vs {meta2['name']})")
            return Response({"error": "Face match inconsistent across frames. Please hold still and try again."}, status=400)
        best_distance = max(dist1, dist2)
        is_real = is_real1 and is_real2
        matched_meta = meta1
    elif meta1:
        best_distance = dist1
        matched_meta = meta1
        is_real = is_real1
    elif meta2:
        best_distance = dist2
        matched_meta = meta2
        is_real = is_real2
        
    if best_distance > MATCH_THRESHOLD:
        print(f"❌ Rejected: Best distance {best_distance:.4f} is above threshold {MATCH_THRESHOLD}")
        return Response({"error": "User Not Found. Face match not confident enough."}, status=404)
        
    # Fetch actual employee object
    matched_employee = Employee.objects.filter(employee_id=matched_meta['employee_id']).first()

    print(f"🏆 FINAL WINNER: {matched_meta['name']} (Dist: {best_distance:.4f})")

    # 5. Handle Spoofing (Only if User Found)
    if not is_real:
        print(f"🚨 DEBUG: Spoofing detected (Liveness Check Failed) for employee {matched_employee.employee_id}!")
        
        # 🚨 Capture Spoofing Attempt
        spoofed_image_b64 = ""
        try:
            if image1_b64:
                spoofed_image_b64 = image1_b64
            elif image1_file:
                image1_file.seek(0)
                file_content = image1_file.read()
                spoofed_image_b64 = base64.b64encode(file_content).decode('utf-8')
        except Exception as e:
            print(f"🚨 DEBUG: Error processing spoofed image for storage: {e}")

        try:
            SpoofingAttempt.objects.create(
                employee_id=matched_employee.employee_id,  # Matched ID
                device_id=device_label,
                image=spoofed_image_b64
            )
            print("✅ DEBUG: SpoofingAttempt record created.")
        except Exception as e:
            print(f"🚨 DEBUG: Error creating SpoofingAttempt record: {e}")

        return Response({"error": "Spoofing detected! Real face required."}, status=400)

    # 6. Mark Attendance (Success)
    mode = request.data.get('mode', 'IN')

    att = EmployeeAttendance.objects.create(
        employee_id=matched_employee.employee_id,
        device_id=device_label,
        attendence_type=mode,
        confidence=best_distance
    )

    print(f"✅ ATTENDANCE SUCCESS: {matched_employee.name} (ID: {matched_employee.employee_id}) | Liveness Verified: {is_real}")

    # Fetch the registered image preview to display on the kiosk
    base64_img = None
    if matched_employee.image_md5:
        try:
            mongo_uri = os.environ.get("GLOBAL_DB_HOST")
            client = MongoClient(mongo_uri)
            
            hr_db_name = os.environ.get("GLOBAL_DB_NAME_HR", "HR")
            global_db_name = os.environ.get("GLOBAL_DB_NAME_GLOBAL", "Global")
            
            fs_hr = gridfs.GridFS(client[hr_db_name])
            fs_global = gridfs.GridFS(client[global_db_name])
            
            file_obj = fs_hr.find_one({"md5": matched_employee.image_md5})
            if not file_obj:
                file_obj = fs_global.find_one({"md5": matched_employee.image_md5})
                
            if file_obj:
                img_bytes = file_obj.read()
                base64_img = f"data:image/jpeg;base64,{base64.b64encode(img_bytes).decode('utf-8')}"
        except Exception as e:
            print(f"🚨 DEBUG: Error fetching registered image for preview: {e}")

    return Response({
        "employee": matched_employee.employee_id,
        "name": matched_employee.name,
        "mode": att.attendence_type,
        "timestamp": att.attendence_time.astimezone(IST),
        "confidence": best_distance,
        "registered_image": base64_img
    }, status=201)


@api_view(['GET'])
@permission_classes([AllowAny])
def attendance_report_with_employee_details(request):
    """
    Get attendance records filtered by date and merged with employee details.
    Example: /api/attendance-report/?from_date=2025-10-01&to_date=2025-10-14
    """
    try:
        # ---- Date Filtering ----
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')

        if not from_date or not to_date:
            now = datetime.now()
            from_date = datetime(now.year, now.month, 1)
            to_date = datetime(now.year, now.month + 1, 1) if now.month < 12 else datetime(now.year + 1, 1, 1)
        else:
            from_date = datetime.strptime(from_date, "%Y-%m-%d")
            to_date = datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1)

        # ---- Fetch Attendance Records ----
        # Fetch from previous day to correctly pair night shifts
        from_date_ext = from_date - timedelta(days=1)
        records = EmployeeAttendance.objects.filter(
            attendence_time__gte=from_date_ext,
            attendence_time__lt=to_date
        ).order_by('employee_id', 'attendence_time')

        if not records.exists():
            return Response([], status=200)

        # ---- Mongo Connection ----
        mongo_uri = os.environ.get("GLOBAL_DB_HOST")
        db_name = os.environ.get("GLOBAL_DB_NAME", "Global")
        client = MongoClient(mongo_uri)
        db = client[db_name]

        profiles = db['backend_diagnostics_profile']
        departments = db['backend_diagnostics_Departments']
        designations = db['backend_diagnostics_Designation']

        # ---- SQL Department Map ----
        from employees.models import Department
        sql_dept_map = {d.name: d.id for d in Department.objects.all()}

        # ---- Create Lookup Maps ----
        dept_map = {
            d.get('department_code'): d.get('department_name')
            for d in departments.find()
        }
        desig_map = {
            d.get('Designation_code'): d.get('designation')
            for d in designations.find({'is_active': True})
        }

        # ---- Department Filtering Resolution ----
        department_filter = request.GET.get('department')
        search_values = []
        if department_filter and department_filter != 'All':
            raw_ids = [d.strip() for d in department_filter.split(',')]
            
            # Resolve SQL IDs to names
            from employees.models import Department as SQLDepartment
            resolved_sql_names = list(SQLDepartment.objects.filter(id__in=[rid for rid in raw_ids if rid.isdigit()]).values_list('name', flat=True))
            
            query_names = raw_ids + resolved_sql_names
            
            # Find Mongo codes for these names/codes
            dept_cursor = departments.find({
                "$or": [
                    {"department_name": {"$in": query_names}},
                    {"department_code": {"$in": query_names}}
                ]
            })
            mongo_codes = [d.get("department_code") for d in dept_cursor]
            
            search_values = list(set(raw_ids + query_names + mongo_codes))

        # ---- Fetch Employee Lookup ----
        employee_map = {}
        # Fetch only employees with face registered from SQL
        face_registered_ids = set(Employee.objects.filter(current_face_encoding__isnull=False).values_list('employee_id', flat=True))
        
        # Profile Query
        profile_query = {}
        if search_values:
            profile_query = {
                "$or": [
                    {"department": {"$in": search_values}},
                    {"department_id": {"$in": search_values}},
                    {"department_name": {"$in": search_values}}
                ]
            }

        # Fetch matching profiles from Mongo
        all_profiles = list(profiles.find(profile_query))
        for emp in all_profiles:
            emp_id = str(emp.get("employeeId"))
            if emp_id not in face_registered_ids:
                continue
                
            dept_code = emp.get("department")
            dept_name = dept_map.get(dept_code, dept_code)
            employee_map[emp_id] = {
                "employeeName": emp.get("employeeName"),
                "department": dept_name,
                "department_id": sql_dept_map.get(dept_name, dept_code),
                "designation": desig_map.get(emp.get("designation"), emp.get("designation")),
            }
            
        # Fallback for SQL registered employees missing from Global DB
        if not search_values or "Unassigned" in search_values:
            registered_sql_emps = Employee.objects.filter(current_face_encoding__isnull=False)
            for sql_emp in registered_sql_emps:
                emp_id_str = str(sql_emp.employee_id)
                if emp_id_str not in employee_map:
                    employee_map[emp_id_str] = {
                        "employeeName": sql_emp.name,
                        "department": "Unassigned",
                        "department_id": None,
                        "designation": "Unassigned"
                    }

        # ---- Prepare Result ----
        result = []
        # (Filtering now handled during population of employee_map above)

        # Date range for iteration (to_date was already +1 day)
        report_dates = []
        curr = from_date
        while curr < to_date:
            report_dates.append(curr.date())
            curr += timedelta(days=1)

        # Group actual records by (employee_id, date) with Shift Date Logic
        records_by_emp_day = {}
        
        current_emp_id = None
        current_shift_date = None
        last_in_time = None
        
        from datetime import time
        noon_time = time(12, 0)
        
        for r in records:
            # Convert to IST
            ist_time = r.attendence_time.astimezone(IST)
            punch_date = ist_time.date()
            
            if current_emp_id != str(r.employee_id):
                current_emp_id = str(r.employee_id)
                current_shift_date = None
                last_in_time = None
                
            punch_type = r.attendence_type
            assigned_date = punch_date
            
            if punch_type == 'IN':
                current_shift_date = punch_date
                last_in_time = ist_time
                assigned_date = current_shift_date
            elif punch_type == 'OUT':
                if current_shift_date and last_in_time:
                    if (ist_time - last_in_time).total_seconds() <= 16 * 3600:
                        assigned_date = current_shift_date
                    else:
                        if ist_time.time() < noon_time:
                            assigned_date = punch_date - timedelta(days=1)
                else:
                    if ist_time.time() < noon_time:
                        assigned_date = punch_date - timedelta(days=1)
            
            key = (str(r.employee_id), assigned_date)
            if key not in records_by_emp_day:
                records_by_emp_day[key] = []
            
            r.attendence_time = ist_time  # Use IST for display/export
            records_by_emp_day[key].append(r)

        # Iterate over ALL employees and ALL dates
        # Sorting by name for consistency
        sorted_emp_ids = sorted(employee_map.keys(), key=lambda eid: employee_map[eid]['employeeName'] or "")

        for emp_id in sorted_emp_ids:
            emp_info = employee_map[emp_id]
            # (Additional in-loop filter not needed as employee_map is already filtered)

            for d in report_dates:
                key = (emp_id, d)
                day_records = records_by_emp_day.get(key, [])
                
                if day_records:
                    for r in day_records:
                        result.append({
                            "employee_id": emp_id,
                            "employee_name": emp_info.get("employeeName", "Unknown"),
                            "department": emp_info.get("department", "N/A"),
                            "department_id": emp_info.get("department_id"),
                            "designation": emp_info.get("designation", "N/A"),
                            "device_id": r.device_id,
                            "attendence_type": r.attendence_type,
                            "attendence_time": r.attendence_time,
                            "confidence": r.confidence,
                        })
                else:
                    # No records for this employee on this day -> Add placeholder
                    # Use start of day as placeholder time (localized to IST)
                    placeholder_time = IST.localize(datetime.combine(d, datetime.min.time()))
                    result.append({
                        "employee_id": emp_id,
                        "employee_name": emp_info.get("employeeName", "Unknown"),
                        "department": emp_info.get("department", "N/A"),
                        "department_id": emp_info.get("department_id"),
                        "designation": emp_info.get("designation", "N/A"),
                        "device_id": "N/A",
                        "attendence_type": "ABSENT",
                        "attendence_time": placeholder_time,
                        "confidence": 0,
                    })

        # ---- Export Support ----
        export_mode = request.GET.get('export')
        if export_mode in ['xlsx', 'csv', 'detailed_xlsx']:
            import pandas as pd
            from django.http import HttpResponse

            if export_mode == 'detailed_xlsx':
                from openpyxl import Workbook
                from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Detailed Attendance"
                
                # Header Styling
                header_font = Font(bold=True, color="000000")
                header_fill = PatternFill(start_color="e2e8f0", end_color="e2e8f0", fill_type="solid")
                alignment_center = Alignment(horizontal="center", vertical="center")
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                # Prepare the Matrix Headers
                base_cols = ['S.No', 'Employee ID', 'Employee Name', 'Department', 'Designation']
                # Row 1: Employee info headers (merged vertically) + Dates (merged horizontally 3 cols each)
                # Row 2: Empty under info headers + In/Out/Total under dates
                
                for i, col_name in enumerate(base_cols, start=1):
                    ws.cell(row=1, column=i, value=col_name).font = header_font
                    ws.cell(row=1, column=i).fill = header_fill
                    ws.cell(row=1, column=i).alignment = alignment_center
                    ws.cell(row=1, column=i).border = thin_border
                    ws.merge_cells(start_row=1, start_column=i, end_row=2, end_column=i)

                curr_col = len(base_cols) + 1
                for d in report_dates:
                    date_str = d.strftime('%d/%m/%Y')
                    ws.cell(row=1, column=curr_col, value=date_str).font = header_font
                    ws.cell(row=1, column=curr_col).fill = header_fill
                    ws.cell(row=1, column=curr_col).alignment = alignment_center
                    ws.cell(row=1, column=curr_col).border = thin_border
                    ws.merge_cells(start_row=1, start_column=curr_col, end_row=1, end_column=curr_col + 2)
                    
                    # Row 2 headers
                    for sub_idx, sub_name in enumerate(['In', 'Out', 'Total']):
                        c = curr_col + sub_idx
                        ws.cell(row=2, column=c, value=sub_name).font = header_font
                        ws.cell(row=2, column=c).fill = header_fill
                        ws.cell(row=2, column=c).alignment = alignment_center
                        ws.cell(row=2, column=c).border = thin_border
                    
                    curr_col += 3

                # Data rows
                # Regroup result data by employee and date
                data_map = {} # (emp_id, date) -> {in, out, total}
                # result is a list of all raw punch/absent recs
                
                # To calculate in/out/total, let's group pulses in result
                for r in result:
                    eid = r.get('employee_id')
                    dt_obj = pd.to_datetime(r.get('attendence_time'))
                    d_key = dt_obj.date()
                    if (eid, d_key) not in data_map:
                        data_map[(eid, d_key)] = {'in': None, 'out': None, 'type': r.get('attendence_type')}
                    
                    entry = data_map[(eid, d_key)]
                    p_type = r.get('attendence_type')
                    if p_type == 'IN':
                        if not entry['in'] or dt_obj < pd.to_datetime(entry['in']):
                            entry['in'] = r.get('attendence_time')
                    elif p_type == 'OUT':
                        if not entry['out'] or dt_obj > pd.to_datetime(entry['out']):
                            entry['out'] = r.get('attendence_time')

                # Write employees (sorted)
                sorted_ids = sorted(employee_map.keys())
                row_idx = 3
                for s_idx, eid in enumerate(sorted_ids, 1):
                    info = employee_map[eid]
                    ws.cell(row=row_idx, column=1, value=s_idx).border = thin_border
                    ws.cell(row=row_idx, column=2, value=eid).border = thin_border
                    ws.cell(row=row_idx, column=3, value=info.get('employeeName', 'Unknown')).border = thin_border
                    ws.cell(row=row_idx, column=4, value=info.get('department', 'N/A')).border = thin_border
                    ws.cell(row=row_idx, column=5, value=info.get('designation', 'N/A')).border = thin_border
                    
                    c_idx = 6
                    for d in report_dates:
                        entry = data_map.get((eid, d), {})
                        in_t = "-"
                        out_t = "-"
                        total_h = 0
                        
                        if entry.get('in'):
                            in_t = pd.to_datetime(entry['in']).strftime('%H:%M:%S')
                        if entry.get('out'):
                            out_t = pd.to_datetime(entry['out']).strftime('%H:%M:%S')
                        
                        if entry.get('in') and entry.get('out'):
                            # Calculate hours (Gross)
                            diff = pd.to_datetime(entry['out']) - pd.to_datetime(entry['in'])
                            total_h = round(diff.total_seconds() / 3600, 2)
                            if total_h < 0: total_h = 0
                        
                        ws.cell(row=row_idx, column=c_idx, value=in_t).border = thin_border
                        ws.cell(row=row_idx, column=c_idx+1, value=out_t).border = thin_border
                        ws.cell(row=row_idx, column=c_idx+2, value=total_h).border = thin_border
                        c_idx += 3
                    
                    row_idx += 1

                # Final response
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename="Detailed_Attendance_Report.xlsx"'
                wb.save(response)
                return response

            # Flat Export 
            df = pd.DataFrame(result)
            if not df.empty:
                # Add Day, Month, Year
                df['Timestamp_obj'] = pd.to_datetime(df['attendence_time'])
                df['Day'] = df['Timestamp_obj'].dt.day
                df['Month'] = df['Timestamp_obj'].dt.month
                df['Year'] = df['Timestamp_obj'].dt.year
                
                df = df[['employee_id', 'employee_name', 'department', 'designation', 'Day', 'Month', 'Year', 'attendence_type', 'attendence_time']]
                df.columns = ['Employee ID', 'Name', 'Department', 'Designation', 'Day', 'Month', 'Year', 'Punch Type', 'Timestamp']
                df['Timestamp'] = pd.to_datetime(df['Timestamp']).dt.strftime('%d/%m/%Y %H:%M:%S')
                df.loc[df['Punch Type'] == 'ABSENT', 'Timestamp'] = '-'

            if export_mode == 'xlsx':
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = r'attachment; filename="Attendance_Report.xlsx"'
                df.to_excel(response, index=False, engine='openpyxl')
                return response
            else:
                response = HttpResponse(content_type='text/csv')
                response['Content-Disposition'] = r'attachment; filename="Attendance_Report.csv"'
                df.to_csv(response, index=False)
                return response

        return Response(result, status=200)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@api_view(['POST'])
def verify_face(request):
    """
    Fast verification endpoint for the first frame in a dual-image sequence.
    Returns success if a confident face match is found, otherwise fails fast.
    """
    image_b64 = request.data.get('image')
    image_file = request.FILES.get('image')
    
    if not image_b64 and not image_file:
        return Response({"error": "Image is required"}, status=400)

    try:
        if image_file:
            enc, is_real = imagefile_to_encoding(image_file)
        else:
            enc, is_real = base64_to_encoding(image_b64)
    except Exception as e:
        print(f"🚨 DEBUG: Error extracting face for verify-face: {e}")
        return Response({"error": "Error processing image"}, status=400)

    if not enc:
        return Response({"error": "No face found in image"}, status=400)

    known_matrix, employee_meta = get_optimized_encodings()
    
    if known_matrix.size == 0:
        return Response({"error": "No registered employees found"}, status=404)

    enc_np = np.array(enc)
    distances = np.linalg.norm(known_matrix - enc_np, axis=1)
    
    best_idx = np.argmin(distances)
    best_distance = float(distances[best_idx])
    matched_meta = employee_meta[best_idx]
    
    MATCH_THRESHOLD = 0.50
    if best_distance > MATCH_THRESHOLD:
        return Response({"error": "User Not Found. Face match not confident enough."}, status=404)
        
    if not is_real:
        # Spoofing detected, fail fast
        return Response({"error": "Spoofing detected (Liveness Check Failed)"}, status=403)
        
    return Response({
        "success": True, 
        "employee_id": matched_meta['employee_id'],
        "name": matched_meta['name']
    }, status=200)


@api_view(['GET'])
@permission_classes([AllowAny])  # Or admin permission
def get_spoofing_attempts(request):
    """
    Get all recorded spoofing attempts.
    Query Params:
    - month: int (1-12)
    - year: int (e.g. 2025)
    """
    month = request.GET.get('month')
    year = request.GET.get('year')
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')

    queryset = SpoofingAttempt.objects.all()

    if from_date_str and to_date_str:
        try:
            start_date = datetime.strptime(from_date_str, "%Y-%m-%d")
            # For 'to_date', we want to include the entire day, so we go to the start of the next day
            end_date_inclusive = datetime.strptime(to_date_str, "%Y-%m-%d")
            end_date = end_date_inclusive + timedelta(days=1)
            queryset = queryset.filter(timestamp__gte=start_date, timestamp__lt=end_date)
        except (ValueError, TypeError):
            pass
    elif month and year:
        try:
            m = int(month)
            y = int(year)
            start_date = datetime(y, m, 1)
            if m == 12:
                end_date = datetime(y + 1, 1, 1)
            else:
                end_date = datetime(y, m + 1, 1)
            
            queryset = queryset.filter(timestamp__gte=start_date, timestamp__lt=end_date)
        except (ValueError, TypeError):
            pass # Invalid month/year, return all


    department = request.GET.get('department')
    if department and department != 'All':
        try:
            mongo_uri = os.environ.get("GLOBAL_DB_HOST")
            db_name = os.environ.get("GLOBAL_DB_NAME", "Global")
            client = MongoClient(mongo_uri)
            db = client[db_name]
            
            raw_values = [d.strip() for d in department.split(',')]
            
            # Resolve numeric IDs to names
            from employees.models import Department
            numeric_ids = [rv for rv in raw_values if rv.isdigit()]
            resolved_names = list(Department.objects.filter(id__in=numeric_ids).values_list('name', flat=True))
            
            dept_query_values = raw_values + resolved_names
            
            # Resolve department codes from names OR treat them as codes directly
            dept_cursor = db['backend_diagnostics_Departments'].find({
                "$or": [
                    {"department_name": {"$in": dept_query_values}},
                    {"department_code": {"$in": dept_query_values}}
                ]
            })
            dept_codes = [d.get("department_code") for d in dept_cursor]
            
            # If nothing found in mongo, fallback to raw values (might be codes already)
            if not dept_codes:
                dept_codes = dept_query_values

            query = {"department": {"$in": dept_codes}}
            dept_profiles = list(db['backend_diagnostics_profile'].find(query, {"employeeId": 1}))
            dept_emp_ids = [str(p["employeeId"]) for p in dept_profiles]

            queryset = queryset.filter(employee_id__in=dept_emp_ids)

        except Exception as e:
            print(f"Error filtering spoofing reports by department: {e}")

    attempts = queryset.order_by('-timestamp')
    data = []
    for attempt in attempts:
        data.append({
            "id": attempt.id,
            "employee_id": attempt.employee_id,
            "device_id": attempt.device_id,
            "timestamp": attempt.timestamp.astimezone(IST),
            "image": attempt.image # Base64 string
        })
    return Response(data, status=200)

@api_view(['POST'])
@permission_classes([AllowAny]) # Should be protected in production
def delete_spoofing_attempts(request):
    """
    Delete spoofing attempts by ID.
    Body: { "ids": [1, 2, 3] } or { "ids": "all", "month": 10, "year": 2025 }
    """
    ids = request.data.get('ids')
    
    if ids == "all":
        # Delete all or filtered all
        # (Optional: Add filtering logic for "delete all in this month" if needed)
        # For now, simplistic "delete selected" vs "delete all" logic based on frontend
        # If user wants "delete all in current filter", frontend sends list of all IDs or we handle filter params here.
        # Let's support list of IDs for safety usually.
        # If the user sends "all", we delete EVERYTHING? That's dangerous. 
        # Better to delete everything matching the current filter if filter params provided?
        # Let's stick to list of IDs for explicit deletion unless requested otherwise.
        # Re-reading request: "table type select all and delete" usually implies client-side select all -> send all IDs.
        pass

    if not ids or not isinstance(ids, list):
         return Response({"error": "Invalid IDs provided"}, status=400)

    try:
        SpoofingAttempt.objects.filter(id__in=ids).delete()
        return Response({"message": "Deleted successfully"}, status=200)
    except Exception as e:
        return Response({"error": str(e)}, status=500)
